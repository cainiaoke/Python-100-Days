#coding=gbk      ###解决编码错误问题
import csv
# with open('E:\\PycharmProjects\\ForTest\\1.csv',newline='') as f:
#     reader = csv.reader(f)
#     for row in reader:
#         print(row)
# with open('E:\\PycharmProjects\\ForTest\\1.csv','a',newline='',encoding='utf-8') as f:
#     writer = csv.writer(f)
#     writer.writerow(['jieshu','34','4'])
# with open('E:\\PycharmProjects\\ForTest\\1.csv',newline='') as f:
#     reader = csv.reader(f)
#     for row in reader:
#         print(row)

# #######发送邮件（纯文本）#######
# import smtplib
# from email.mime.text import MIMEText
# from email.header import Header
# mail_host='smtp.isoftstone.com'
# mail_user='kehana'
# mail_pass='1QAZ@2WSX'
# sender='kehana@isoftstone.com'
# receivers=['cainiaoke@163.com']
# content=input('请输入邮件正文：')
# message=MIMEText(content,'plain','utf-8')   #设置正文格式
# subject=input('请输入邮件主题：')
# message['Subject']=Header(subject,'utf-8')    #设置主题格式
# message['From']=sender
# message['To']=receivers[0]
# try:
#     smtpObj=smtplib.SMTP()
#     smtpObj.connect(mail_host,25)
#     ##有些邮箱开启了SSl认证，所以以上两步要替换为
#     # smtpObj=smtplib.SMTP_SSL(mail_host)
#     smtpObj.login(mail_user,mail_pass)
#     smtpObj.sendmail(sender,receivers,message.as_string())
#     smtpObj.quit()
#     print('success')
# except smtplib.SMTPException as e:
#     print('error',e)

# #######发送邮件（带附件）#######
# import smtplib
# from email.mime.text import MIMEText
# from email.mime.multipart import MIMEMultipart
# from email.mime.image import MIMEImage
# from email.header import Header
# #设置登录以及服务器信息
# mail_host='smtp.isoftstone.com'
# mail_user='kehana'
# mail_pass='1QAZ@2WSX'
# sender='kehana@isoftstone.com'
# receivers='cainiaoke@163.com'
# #设置email信息，添加一个MIME用来处理附件信息
# message=MIMEMultipart()
# message['From']=sender
# message['To']=receivers
# biaoti=input('请输入邮件标题：')
# message['Subject']=biaoti
# #添加正文和2个附件
# with open('E:\\PycharmProjects\\ForTest\\zhihu.html','r',encoding='utf-8') as h:  #encoding='utf-8'防止出现编码问题
#     content=h.read()
# Content=MIMEText(content,'html','utf-8')
# with open('E:\\PycharmProjects\\ForTest\\tmp1.txt','r') as t:
#     content2=t.read()
# fujian1=MIMEText(content2,'plain','utf-8')
# fujian1['Content-Type']='application/octet-stream'
# fujian1['Content-disposition']='attachment;filename="tmp.txt"'
# with open('E:\\PycharmProjects\\ForTest\\1.JPG','rb') as j:
#     picture=MIMEImage(j.read())
#     picture['Content-Type']='application/octet-stream'
#     picture['Content-disposition']='attachment;filename="1.jpg"'
# message.attach(Content)
# message.attach(fujian1)
# message.attach(picture)
# try:
#     smtpObj=smtplib.SMTP()
#     smtpObj.connect(mail_host,25)
#     smtpObj.login(mail_user,mail_pass)
#     smtpObj.sendmail(sender,receivers,message.as_string())
#     print('success')
#     smtpObj.quit()
# except smtplib.SMTPException as e:
#     print('error',e)

# #检查是否安装python模块。第一种方法：
# try:
#     import requests
# except ImportError:
#     print("模块不存在")
# #检查是否安装python模块。第二种方法：
# name=input("请输入模块名称：")
# try:
#     import name
# except ImportError:
#     print("模块不存在")
# #检查是否安装python模块。第三种方法：
# name=input("请输入模块名称：")
# code="import "+name
# try:
#     exec(code)
# except ImportError:
#     print("模块不存在")
# #检查是否安装python模块。第四种方法：
# import os
# name=input("请输入模块名：")
# all=os.popen("pip list").read()
# if name in all:
#     print("模块已安装")
# else:
#     print("模块未安装")

# import requests
# res=requests.get('https://localprod.pandateacher.com/python-manuscript/crawler-html/sanguo.md')
# print(res.encoding)  #查看网页的编码格式
# # res.encoding='gbk'   #修改网页的编码格式
# novel=res.text   #text表示转化为文本形式
# #novel=res.status_code  #status_code表示检查请求是否成功，200：成功，100：继续提出请求，403：禁止访问，503：服务器不可用
# #print(novel[:800])  #打印前800行
# # novel1=novel[:100]
# # print(novel1)
# S=open('《三国演义》.txt','a')
# S.write(novel)
# S.close()

# import requests
# res=requests.get('https://localprod.pandateacher.com/python-manuscript/crawler-html/sanguo.md')
# print(res.status_code)
# novel=res.text
# print(novel[:200])
# s=open("《三国演义》.txt","a+")
# s.write(novel)
# s.close()

# import requests
# res=requests.get('https://res.pandateacher.com/2019-01-12-15-29-33.png')
# p=res.content   #二进制图片用content
# file=open("图片.jpg","wb")   #t二进制用wb
# file.write(p)
# file.close()

# import requests
# huoqu = requests.get ( 'https://localprod.pandateacher.com/python-manuscript/crawler-html/spider-men5.0.html' )
# print ( huoqu.text )
# from bs4 import BeautifulSoup
# jiexi=BeautifulSoup(huoqu.text,'html.parser')
# # print(jiexi)
# tiqus=jiexi.find_all(class_='books')
# print(tiqus)
# for tiqu in tiqus:
#     kind=tiqu.find('h2')
#     title=tiqu.find(class_="title")
#     info=tiqu.find(class_="info")
#     print(kind,'\n',title,'\n',info,'\n')
    # print(' ',kind.text,'\n',title['href'],'\n',title.text,'\n',info.text,'\n')

# import requests
# huoqu=requests.get('https://wordpress-edu-3autumn.localprod.oc.forchange.cn/all-about-the-future_04/')
# # print(huoqu.text)
# from bs4 import BeautifulSoup
# jiexi=BeautifulSoup(huoqu.text,'html.parser')
# tiqus=jiexi.find_all(class_='comment-content')
# print(tiqus)
# for tiqu in tiqus:
#     print(tiqu.text)

# import requests
# from bs4 import BeautifulSoup
# url=('http://books.toscrape.com/')
# huoqu=requests.get(url)
# jiexi=BeautifulSoup(huoqu.text,"html.parser")
# tiqus=jiexi.find('ul',class_="nav").find('ul').find_all('li')
# # print(tiqus)
# for tiqu in tiqus:
#     name=tiqu.find('a')
#     print(name.text.strip()) #strip()表示去掉空格和空行

# #打印出书名，star，价格
# import requests
# from bs4 import BeautifulSoup
# url=('http://books.toscrape.com/')
# huoqu=requests.get(url)
# jiexi=BeautifulSoup(huoqu.text,'html.parser')
# tiqus=jiexi.find_all(class_="col-xs-6 col-sm-4 col-md-3 col-lg-3")
# # print(tiqus)
# for tiqu in tiqus:
#     name=tiqu.find('h3').find('a')
#     price=tiqu.find(class_="price_color")
#     star=tiqu.find('p',class_="star-rating")
#     print(name['title'],price.text,star['class'][1])

# import requests
# from bs4 import BeautifulSoup
# huoqu=requests.get('https://wordpress-edu-3autumn.localprod.oc.forchange.cn/')
# jiexi=BeautifulSoup(huoqu.text,'html.parser')
# tiqus=jiexi.find_all('header',class_="entry-header")
# # print(tiqus)
# for tiqu in tiqus:
#     shijian=tiqu.find('time')
#     biaoti=tiqu.find('h2')
#     lianjie=tiqu.find('a')
#     print(biaoti.text,'\n',lianjie['href'],'\n',shijian.text,'\n')

##把豆瓣TOP250里面的 序号/电影名/评分/推荐语/链接 都爬取下来
# import requests
# from bs4 import BeautifulSoup
# #requests.get报错418，404，ctrl+shift+i进到Network,ALL,随便点一个Name里面的内容右边会出来Requetst Headers,翻到最下面有user-agent:赋值给headers，注意{ }里面有两个单引号
# headers={'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
# url='https://movie.douban.com/top250?start=0&filter='
# huoqu=requests.get(url,headers=headers)
# jiexi=BeautifulSoup(huoqu.text,'html.parser')
# tiqus=jiexi.find('ol',class_="grid_view").find_all('li')
# #print(tiqus)
# for tiqu in tiqus:
#     xuhao=tiqu.find('em',class_="").text
    # name=tiqu.find_all('span',class_="title")
    # pingfen=tiqu.find('span',class_="rating_num")
    # tuijianyu=tiqu.find('span',class_="inq")
    # lianjie=tiqu.find('a',class_="")
    # print(' ',xuhao.text,'\n',name.text,'\n',tuijianyu.text,'\n',pingfen.text,'\n',lianjie['href'],'\n')
    # print(xuhao.text)

## 菜名、原材料、详细烹饪流程的URL
# import requests
# from bs4 import BeautifulSoup
# url='http://www.xiachufang.com'
# headers={'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
# huoqu=requests.get(url+'/explore/',headers=headers)
# jiexi=BeautifulSoup(huoqu.text,'html.parser')
# tiqus=jiexi.find(class_="normal-recipe-list").find_all('li')
# for tiqu in tiqus:
#     name=tiqu.find('img')
#     yuanliao=tiqu.find(class_="ing ellipsis")
#     lianjie1=tiqu.find(class_="name").find('a')
#     lianjie2=lianjie1['href']
#     # print(name['alt'])
#     # print(yuanliao.text.strip())
#     print(' 菜  名：'+name['alt'],'\n','原材料：'+yuanliao.text.strip(),'\n','链  接：',url+lianjie2,'\n')

##序号/电影名/评分/推荐语/链接
# import requests
# from bs4 import BeautifulSoup
# for i in range(10):  ##爬取所有页内容
#     url='https://movie.douban.com/top250?start='+str(i*25)+'&filter='
#     headers={'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
#     huoqu=requests.get(url,headers=headers)
#     jiexi=BeautifulSoup(huoqu.text,'html.parser')
#     tiqus=jiexi.find(class_="grid_view").find_all('li')
#     for tiqu in tiqus:
#         xuhao=tiqu.find(class_="").text
#         name=tiqu.find(class_="title").text
#         pingfen=tiqu.find(class_="rating_num").text
#         lianjie=tiqu.find('a')['href']
#         if tiqu.find('span',class_="inq") != None:  ##判断推荐语是否为空
#             tuijianyu=tiqu.find('span',class_="inq").text
#             print(' 序  号：'+str(xuhao),'\n','电影名：'+name,'\n','评  分：'+str(pingfen),'\n','推荐语：'+tuijianyu,'\n','链  接：'+str(lianjie),'\n')
#         else:
#             print(' 序  号：'+str(xuhao),'\n','电影名：'+name,'\n','评  分：'+str(pingfen),'\n','链  接：'+str(lianjie),'\n')

# import requests
# from bs4 import BeautifulSoup
# from urllib.request import quote   ###中文转化为URL
# a=input('请输入电影名：')
# b=a.encode('gbk')
# url1='http://s.ygdy8.com/plus/s0.php?typeid=1&keyword='+quote(b)
# huoqu1=requests.get(url1)
# huoqu1.encoding='gbk'
# jiexi1=BeautifulSoup(huoqu1.text,'html.parser')
# try:
#     tiqu1=jiexi1.find(class_="co_content8").find('table').find('a')['href']
#     # print(tiqu1)
#     if tiqu1:
#         url2='https://www.ygdy8.com'+tiqu1
#         # print(url2)
#         huoqu2=requests.get(url2)
#         huoqu2.encoding='gbk'
#         jiexi2=BeautifulSoup(huoqu2.text,'html.parser')
#         lianjie=jiexi2.find('div',id="Zoom").find('span').find('table').find('a')['href']
#         print(lianjie)
#     else:
#         print('没有 《'+a+'》 的下载链接')
# except:
#     print('没有找到电影')

#encoding=utf-8
# from urllib.request import quote
# from bs4 import BeautifulSoup
# import requests
# # name=input('请输入歌手名字：')
# # name1=name.encode('utf-8')
# # url='https://y.qq.com/portal/search.html#page=1&searchid=1&remoteplace=txt.yqq.top&t=song&w='+quote(name)
# # huoqu=requests.get('https://c.y.qq.com/soso/fcgi-bin/client_search_cp?ct=24&qqmusic_ver=1298&new_json=1&remoteplace=txt.yqq.song&searchid=68293098210396354&t=0&aggr=1&cr=1&catZhida=1&lossless=0&flag_qc=0&p=1&n=10&w=%E5%91%A8%E6%9D%B0%E4%BC%A6&g_tk_new_20200303=5381&g_tk=5381&loginUin=0&hostUin=0&format=json&inCharset=utf8&outCharset=utf-8&notice=0&platform=yqq.json&needNewCode=0')
# # tiqus=huoqu.json()
# # tiqu_list=tiqus['data']['song']['list']
# # print(tiqu_list)
# # for tiqu1 in tiqu_list:
# #     print('歌曲名：'+tiqu1['name'])
# #     print('专辑名：'+tiqu1['album']['name'])
# #     print('播放链接：'+'https://y.qq.com/n/yqq/song/'+tiqu1['mid']+'.html')
# #     url1='https://y.qq.com/n/yqq/song/'+tiqu1['mid']+'.html'
# #     huoqu1=requests.get(url1)
# #     jiexi=BeautifulSoup(huoqu1.text,'html.parser')
# #     tiqu2=jiexi.find(class_="lyric__cont_box")
# #     print('歌词：'+tiqu2.text)

# import json
# a=[1,2,3,4]
# b=json.dumps(a)   #dumps是将列表或字典转化为json字符串
# print(type(a))
# print(type(b))
# c=json.loads(b)    #loads是将json字符串转化为列表或者字典
# print(c)
# print(type(c))

# ###爬取歌手的歌曲信息
# import requests
# from urllib.request import quote   ####quote是将非ASCII编码（ASCII编码指部分数字，字母，特殊符号）的字符转化为URL编码
# a=input('请输入歌手名字：')
# # geshou=quote(a)
# for i in range(5):
#     params={
#     'ct':'24',
#     'qqmusic_ver':'1298',
#     'new_json':'1',
#     'remoteplace':'txt.yqq.song',
#     'searchid':'56587706012726378',
#     't':'0',
#     'aggr':'1',
#     'cr':'1',
#     'catZhida':'1',
#     'lossless':'0',
#     'flag_qc':'0',
#     'p':str(i+1),
#     'n':'10',
#     'w':a,
#     'g_tk_new_20200303':'5381',
#     'g_tk':'5381',
#     'loginUin':'0',
#     'hostUin':'0',
#     'format':'json',
#     'inCharset':'utf8',
#     'outCharset':'utf-8',
#     'notice':'0',
#     'platform':'yqq.json',
#     'needNewCode':'0',
#     }
#     url='https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
#     huoqu=requests.get(url,params=params)
#     tiqus=huoqu.json()
#     tiqu_list=tiqus['data']['song']['list']
#     for music in tiqu_list:
#         print('歌曲名：'+music['name'])
#         print('专辑名：'+music['album']['name'])
#         print('歌曲链接：'+'https://y.qq.com/n/yqq/song/'+music['mid']+'.html''\n\n')

# ###爬取周杰伦歌词
# import requests
# import json
# # 引用requests,json模块
# url = 'https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
# headers = {
#     'user-agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36'
#     # 标记了请求从什么设备，什么浏览器上发出
#     }
# for x in range(5):
#     params = {
#     'ct':'24',
#     'qqmusic_ver': '1298',
#     'new_json':'1',
#     'remoteplace':'sizer.yqq.lyric_next',
#     'searchid':'94267071827046963',
#     'aggr':'1',
#     'cr':'1',
#     'catZhida':'1',
#     'lossless':'0',
#     'sem':'1',
#     't':'7',
#     'p':str(x+1),
#     'n':'10',
#     'w':'周杰伦',
#     'g_tk':'1714057807',
#     'loginUin':'0',
#     'hostUin':'0',
#     'format':'json',
#     'inCharset':'utf8',
#     'outCharset':'utf-8',
#     'notice':'0',
#     'platform':'yqq.json',
#     'needNewCode':'0'
#     }
#     res = requests.get(url, params = params)
#     # print(params)
#     # 下载该网页，赋值给res
#     jsonres = json.loads(res.text)
#     #使用json来解析res.text
#     list_lyric = jsonres['data']['lyric']['list']
#     #一层一层地取字典，获取歌词的列表
#     for lyric in list_lyric:
#     #lyric是一个列表，x是它里面的元素
#         print(lyric['content'])
#     # 以content为键，查找歌词

###快递100查快递
# import requests
# gs=input('请输入快递公司名称：')
# dh=input('请输入快递单号：')
# mobile=input('请输入手机后四位：')
# headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
# url='https://www.kuaidi100.com/query?type='+gs+'&postid='+dh+'&temp=0.8823894021412677&phone='+mobile
# print(url)
# huoqu=requests.get(url,headers=headers)
# tiqus=huoqu.json()
# print(tiqus['data'][0]['context'])
# # print(huoqu.text)

####向csv文件中写入数据
# import csv
# csv_file=open('demo.csv','a',newline='',encoding='gbk')
# xieru=csv.writer(csv_file)
# xieru.writerow(['阿甘正传','10'])
# xieru.writerow(['当幸福来敲门','9'])
# csv_file.close()
####读取csv文件中的数据
# csv_file=open('demo.csv','r',newline='')
# du=csv.reader(csv_file)
# for row in du:
#     print(row)
# csv_file.close()

####向xlsx表格写入数据
import openpyxl
# demo1=openpyxl.Workbook()
# sheet=demo1.active
# sheet.title='first'
# row1=[['漫威','日本','中国'],['韩国','美国','欧洲']]
# for i in row1:
#     sheet.append(i)
# demo1.save('demo1.xlsx')
####读取xlsx表格里面的数据
# demo1=openpyxl.load_workbook('demo1.xlsx')
# sheetname=demo1.sheetnames
# print(sheetname)
# sheet=demo1['first']
# A1_cell=sheet['A1']
# A1_value=A1_cell.value
# print(A1_value)

# ###爬取歌手的歌曲信息并保存至excel
# import requests,openpyxl
# from urllib.request import quote   ####quote是将非ASCII编码（ASCII编码指部分数字，字母，特殊符号）的字符转化为URL编码
#
# excel=openpyxl.Workbook()  #创建一个excel
# sheet=excel.active   #创建一个活动表
# sheet.title='song_list'   #给活动表命名
# sheet['A1']='歌曲名'
# sheet['B1']='专辑名'
# sheet['C1']='歌曲链接'
# sheet['D1']='歌曲时长（s）'
#
# a=input('请输入歌手名字：')
# print('---------------------''\n')
# # geshou=quote(a)
# for i in range(5):
#     params={
#     'ct':'24',
#     'qqmusic_ver':'1298',
#     'new_json':'1',
#     'remoteplace':'txt.yqq.song',
#     'searchid':'56587706012726378',
#     't':'0',
#     'aggr':'1',
#     'cr':'1',
#     'catZhida':'1',
#     'lossless':'0',
#     'flag_qc':'0',
#     'p':str(i+1),
#     'n':'10',
#     'w':a,
#     'g_tk_new_20200303':'5381',
#     'g_tk':'5381',
#     'loginUin':'0',
#     'hostUin':'0',
#     'format':'json',
#     'inCharset':'utf8',
#     'outCharset':'utf-8',
#     'notice':'0',
#     'platform':'yqq.json',
#     'needNewCode':'0',
#     }
#     url='https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
#     huoqu=requests.get(url,params=params)
#     tiqus=huoqu.json()
#     tiqu_list=tiqus['data']['song']['list']
#     for music in tiqu_list:
#         # print('歌曲名：'+music['name'])
#         # print('专辑名：'+music['album']['name'])
#         # print('歌曲链接：'+'https://y.qq.com/n/yqq/song/'+music['mid']+'.html')
#         # print('歌曲时长：'+str(music['interval'])+'s''\n\n')
#         name=music['name']
#         zhuanji=music['album']['name']
#         link='https://y.qq.com/n/yqq/song/'+str(music['mid'])+'.html'
#         time=music['interval']
#         sheet.append([name,zhuanji,link,time])   #将查出的内容append到sheet中
#         print('歌曲名：'+name+'\n'+'专辑名：'+zhuanji+'\n'+'歌曲链接：'+link+'\n'+'歌曲时长（s）'+str(time)+'\n')
# excel.save('歌曲清单.xlsx')  #保存并命名excel表格

####爬取豆瓣TOP250影片信息并保存为excel
# #encoding=utf-8
# import requests,openpyxl
# from bs4 import BeautifulSoup
# from urllib.request import quote
#
# excel=openpyxl.Workbook()
# sheet=excel.active
# sheet.title='list'
# sheet['A1']='序号'
# sheet['B1']='影片名'
# sheet['C1']='评分'
# sheet['D1']='评语'
# sheet['E1']='链接'
#
# for i in range(10):
#     url='https://movie.douban.com/top250?start='+str(i*25)+'&filter='
#     headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
#     huoqu=requests.get(url,headers=headers)
#     huoqu.encoding='utf-8'
#     jiexi=BeautifulSoup(huoqu.text,'html.parser')
#     tiqus=jiexi.find(class_="grid_view").find_all('li')
#     for tiqu in tiqus:
#         xuhao=tiqu.find('em').text
#         name=tiqu.find(class_="title").text
#         pingfen=tiqu.find(class_="rating_num").text
#         lianjie=tiqu.find('a')['href']
#         if tiqu.find('span',class_="inq") != None:
#             pingyu=tiqu.find('span',class_="inq").text
#             sheet.append([xuhao,name,pingfen,pingyu,lianjie])
#             print(' 序  号：'+str(xuhao),'\n','影片名：'+name,'\n','评  分：'+str(pingfen),'\n','评  语：'+pingyu,'\n','链  接：'+str(lianjie),'\n')
#         else:
#             sheet.append([xuhao,name,pingfen,'无评语',lianjie])
#             print(' 序  号：'+str(xuhao),'\n','影片名：'+name,'\n','评  分：'+str(pingfen),'\n','评  语：无评语','\n','链  接：'+str(lianjie),'\n')
#
# excel.save('豆瓣TOP250.xlsx')

###已知用户密码登录某网站
# import requests
# from bs4 import BeautifulSoup
# url='https://wordpress-edu-3autumn.localprod.oc.forchange.cn/wp-login.php'
# headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
# data={
#     'og': 'heheh',
#     'pwd': 'abcd1234',
#     'wp-submit': '登录',
#     'redirect_to': 'https://wordpress-edu-3autumn.localprod.oc.forchange.cn',
#     'testcookie': '1'
# }
# login=requests.post(url,headers=headers,data=data)   ##需要输入用户名密码的时候用post请求
# print(login)

#selenium中方法的用法
# # 以下方法都可以从网页中提取出'你好，蜘蛛侠！'这段文字
#
# find_element_by_tag_name：通过元素的名称选择
# # 如<h1>你好，蜘蛛侠！</h1>
# # 可以使用find_element_by_tag_name('h1')
#
# find_element_by_class_name：通过元素的class属性选择
# # 如<h1 class="title">你好，蜘蛛侠！</h1>
# # 可以使用find_element_by_class_name('title')
#
# find_element_by_id：通过元素的id选择
# # 如<h1 id="title">你好，蜘蛛侠！</h1>
# # 可以使用find_element_by_id('title')
#
# find_element_by_name：通过元素的name属性选择
# # 如<h1 name="hello">你好，蜘蛛侠！</h1>
# # 可以使用find_element_by_name('hello')
#
# #以下两个方法可以提取出超链接
#
# find_element_by_link_text：通过链接文本获取超链接
# # 如<a href="spidermen.html">你好，蜘蛛侠！</a>
# # 可以使用find_element_by_link_text('你好，蜘蛛侠！')
#
# find_element_by_partial_link_text：通过链接的部分文本获取超链接
# # 如<a href="https://localprod.pandateacher.com/python-manuscript/hello-spiderman/">你好，蜘蛛侠！</a>
# # 可以使用find_element_by_partial_link_text('你好')

# def pri(n):   #def定义一个函数，(n)表示函数的参数
#     print(n)    #此函数执行的过程
# pri('韩可')   #调用pri函数并加入参数

# print('打扫卫生步骤如下')
# def step_1():
#     print('取出扫帚')
# def step_2():
#     print('开始打扫卫生')
# def step_3():
#     print('放回扫帚')
# step_1()
# step_2()
# step_3()

# class Cat:     #创建一个类Cat
#     def __init__(self,new_name,new_age):    #定义一个初始化方法，会自动调用，不需要手动
#         self.name=new_name
#         self.age=new_age
#     def eat(self):   #定义一个普通方法
#         print('猫在吃鱼')
#     def drink(self):   #定义一个普通方法
#         print('猫在喝百事')
#     def introduce(self):   #定义一个普通方法
#         print('%s今年%d岁'%(self.name,self.age))
# Tom=Cat('汤姆',18)   #实例化一个对象Tom，并调用Cat类，初始化方法中的self就变成了Tom,初始化方法中的new_name就变成了'汤姆',new_age就变成了18
# Tom.eat()    #对象Tom调用Cat类里面的方法
# Tom.drink()   #对象Tom调用Cat类里面的方法
# Tom.introduce()    #对象Tom调用Cat类里面的方法

# class Student:    #创建一个类Student
#     def __init__(self):    #初始化方法没有加入参数
#         self.name=None
#         self.age=None
#     def introduce(self):
#         print(self.name)
# student=Student()    #实例化一个对象，初始化方法中的self就变成了student
# student.name='汤姆'
# student.age=18
# student.introduce()
#
# class Student1:
#     def __init__(self,new_name,new_age):  #初始化方法中加入两个参数
#         self.name=new_name
#         self.age=new_age
#     def introduce(self):
#         print(self.name,'\n',self.age)
# student1=Student1('汤姆1',19)
# student1.introduce()

# ###已知用户密码登录某网站，生成cookies
# import requests
# url='https://wordpress-edu-3autumn.localprod.oc.forchange.cn/wp-login.php'
# headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
# data={
#     'og': 'heheh',
#     'pwd': 'abcd1234',
#     'wp-submit': '登录',
#     'redirect_to': 'https://wordpress-edu-3autumn.localprod.oc.forchange.cn',
#     'testcookie': '1'
# }
# login=requests.post(url,headers=headers,data=data)   ##需要输入用户名密码的时候用post请求
# cookies=login.cookies  ###调用requests对象login的cookies属性，并复赋值给cookies
# #####利用上面的cookies进入博客发表评论
# url_1='https://wordpress-edu-3autumn.localprod.oc.forchange.cn/wp-content/uploads/2019/01/cropped-cobweb-1698801_1920-1-32x32.jpg'
# data_1={
#     'comment': input('请输入评论内容：'),
#     'submit': '发表评论',
#     'comment_post_ID': '13',
#     'comment_parent': '0'
# }
# comment=requests.post(url_1,headers=headers,cookies=cookies,data=data_1)  ###调用各种参数变量
# print(comment.status_code)

# import requests,json
# session=requests.session()
# url='https://wordpress-edu-3autumn.localprod.oc.forchange.cn/wp-login.php'
# headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
# data={
#     'log': input('请输入用户名：'),
#     'pwd': input('请输入密码：'),
#     'wp-submit': '登录',
#     'redirect_to': 'https://wordpress-edu-3autumn.localprod.oc.forchange.cn',
#     'testcookie': '1'
# }
# session.post(url,headers=headers,data=data)
# print(session.cookies)
# cookies_dic=requests.utils.dict_from_cookiejar(session.cookies)   ####将cookies转化为字典
# print(cookies_dic)
# cookies_str=json.dumps(cookies_dic)   ####将字典转化为json字符串
# print(cookies_str)
# f=open('cookies.txt','w')
# f.write(cookies_str)
# f.close()
# url_1='https://wordpress-edu-3autumn.localprod.oc.forchange.cn/wp-comments-post.php'
# data_1={
#     'comment': input('请输入评论：'),
#     'submit': '发表评论',
#     'comment_post_ID': '15',
#     'comment_parent': '0'
# }
# comment=session.post(url_1,headers=headers,data=data_1)
# print(comment.status_code)

###读取本地cookies登录博客并评论
# import requests,json
# session=requests.session()
# headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
# try:
#     cookies_txt=open('cookies.txt','r')
#     cookies_dic=json.loads(cookies_txt.read())
#     cookies=requests.utils.cookiejar_from_dict(cookies_dic)
#     session.cookies=cookies
# except FileNotFoundError:
#     url='https://wordpress-edu-3autumn.localprod.oc.forchange.cn/wp-login.php'
#     data={
#         'log': input('请输入用户名：'),
#         'pwd': input('请输入密码：'),
#         'wp-submit': '登录',
#         'redirect_to': 'https://wordpress-edu-3autumn.localprod.oc.forchange.cn',
#         'testcookie': '1'
#     }
#     session.post(url,headers=headers,data=data)
#     cookies_dic=requests.utils.dict_from_cookiejar(session.cookies)
#     cookies_str=json.dumps(cookies_dic)
#     f=open('cookies.txt','w')
#     f.write(cookies_str)
#     f.close()
# url_1='https://wordpress-edu-3autumn.localprod.oc.forchange.cn/wp-comments-post.php'
# data_1={
#     'comment': input('请输入评论：'),
#     'submit': '发表评论',
#     'comment_post_ID': '15',
#     'comment_parent': '0'
# }
# comment=session.post(url_1,headers=headers,data=data_1)
# print(comment.status_code)

# #####翻译小程序
# import requests
# import json
# from tkinter import Tk,Button,Entry,Label,Text,END
#
# class YouDaoFanyi(object):
#     def __init__(self):
#         pass
#     def crawl(self,word):
#         url='http://fanyi.youdao.com/translate?smartresult=dict&smartresult=rule'
#         #使用post需要一个链接
#         data={'i': word,
#               'from': 'AUTO',
#               'to': 'AUTO',
#               'smartresult': 'dict',
#               'client': 'fanyideskweb',
#               'doctype': 'json',
#               'version': '2.1',
#               'keyfrom': 'fanyi.web',
#               'action': 'FY_BY_REALTIME',
#               'typoResult': 'false'}
#         #将需要post的内容，以字典的形式记录在data内。
#         r = requests.post(url, data)
#         #post需要输入两个参数，一个是刚才的链接，一个是data，返回的是一个Response对象
#         answer=json.loads(r.text)
#         #你可以自己尝试print一下r.text的内容，然后再阅读下面的代码。
#         result = answer['translateResult'][0][0]['tgt']
#         return result
#
#
#
# class Application(object):
#     def __init__(self):
#         self.window = Tk()
#         self.fanyi = YouDaoFanyi()
#
#
#         self.window.title(u'我的翻译')
#         #设置窗口大小和位置
#         self.window.geometry('310x370+500+300')
#         self.window.minsize(310,370)
#         self.window.maxsize(310,370)
#         #创建一个文本框
#         #self.entry = Entry(self.window)
#         #self.entry.place(x=10,y=10,width=200,height=25)
#         #self.entry.bind("<Key-Return>",self.submit1)
#         self.result_text1 = Text(self.window,background = 'azure')
#         # 喜欢什么背景色就在这里面找哦，但是有色差，得多试试：http://www.science.smith.edu/dftwiki/index.php/Color_Charts_for_TKinter
#         self.result_text1.place(x = 10,y = 5,width = 285,height = 155)
#         self.result_text1.bind("<Key-Return>",self.submit1)
#
#         #创建一个按钮
#         #为按钮添加事件
#         self.submit_btn = Button(self.window,text=u'翻译',command=self.submit)
#         self.submit_btn.place(x=205,y=165,width=35,height=25)
#         self.submit_btn2 = Button(self.window,text=u'清空',command = self.clean)
#         self.submit_btn2.place(x=250,y=165,width=35,height=25)
#
#         #翻译结果标题
#         self.title_label = Label(self.window,text=u'翻译结果:')
#         self.title_label.place(x=10,y=165)
#         #翻译结果
#
#         self.result_text = Text(self.window,background = 'light cyan')
#         self.result_text.place(x = 10,y = 190,width = 285,height = 165)
#         #回车翻译
#     def submit1(self,event):
#         #从输入框获取用户输入的值
#         content = self.result_text1.get(0.0,END).strip().replace("\n"," ")
#         #把这个值传送给服务器进行翻译
#
#         result = self.fanyi.crawl(content)
#         #将结果显示在窗口中的文本框中
#
#         self.result_text.delete(0.0,END)
#         self.result_text.insert(END,result)
#
#         #print(content)
#
#     def submit(self):
#         #从输入框获取用户输入的值
#         content = self.result_text1.get(0.0,END).strip().replace("\n"," ")
#         #把这个值传送给服务器进行翻译
#
#         result = self.fanyi.crawl(content)
#         #将结果显示在窗口中的文本框中
#
#         self.result_text.delete(0.0,END)
#         self.result_text.insert(END,result)
#         print(content)
#     #清空文本域中的内容
#     def clean(self):
#         self.result_text1.delete(0.0,END)
#         self.result_text.delete(0.0,END)
#
#     def run(self):
#         self.window.mainloop()
#
#
# if __name__=="__main__":
#     app = Application()
#     app.run()

###本地Chrome浏览器设置方法
# import time
# from bs4 import BeautifulSoup
# from selenium import webdriver    #导入webdirver
# driver=webdriver.Chrome()   #设置搜索引擎为谷歌
# driver.get('https://localprod.pandateacher.com/python-manuscript/hello-spiderman/')  #打开网页
# time.sleep(2)
# # label=driver.find_element_by_tag_name('label')  ###解析并提取出label标签中的字符串
# # print(label.text)
# # labels=driver.find_elements_by_tag_name('label')
# # print(type(labels))
# # for i in labels:
# #     print(i)
# teacher=driver.find_element_by_id('teacher')
# teacher.send_keys('必须是韩可啊')   ###在输入框输入内容
# time.sleep(3)
# assistant=driver.find_element_by_name('assistant')
# assistant.send_keys('都喜欢')
# time.sleep(3)
# button=driver.find_element_by_cldsadass_name('sub')   #找到提交按钮
# button.click()      #点击提交操作
# time.sleep(5)
# driver.close()

# ##爬取某首歌的热门评论前两页,并采取静默模式
# import time
# from selenium import webdriver
# from selenium.webdriver.chrome.options import Options  #从options模块调用Options类
# chrome_options=Options()  #实例化Options对象
# chrome_options.add_argument('--headless')  #吧谷歌浏览器设置为静默模式不显示出来
# driver=webdriver.Chrome(options=chrome_options)   #设置谷歌浏览器，并在后台运行
# driver.get('https://y.qq.com/n/yqq/song/000xdZuV2LcQ19.html')   #获取歌曲链接
# time.sleep(2)
# button=driver.find_element_by_class_name('js_get_more_hot')    #点击“点击加载更多”
# button.click()    #执行点击操作
# time.sleep(2)
# comments=driver.find_element_by_class_name('js_hot_list').find_elements_by_class_name('js_cmt_li')
# ###以上两个find是为了找到评论的所有内容
# print(len(comments))   #打印有多少条评论
# for comment in comments:
#     sweet=comment.find_element_by_tag_name('p')  #获取每一条具体的评论内容
#     print('评论：%s\n ---\n'%sweet.text)
# driver.close()   #关闭浏览器

#调用接口测试页面
# from flask import Flask,request
# app=Flask(__name__)
# @app.route('/helloo')
# def index():
#     name=request.args.get('name')
#     return '你好，'+name
# if __name__=='__main__':
#     app.run()

# import re
# text='asdfsahanke,hankekekke,hankehanke'   #字符串
# pattern=r'hanke'    #表示被匹配的内容
# pattern=r','  #被匹配的内容
# repl='、'   #替换内容
# print('search:',re.search(pattern,text).group())  #search从字符串中任意位置开始匹配
# print('match:',re.match(pattern,text).group())   #match从头开始匹配
# print('findall:',re.fullmatch(pattern,text).group())  #fullmatch需要完全匹配
# print('dinfall:',re.findall(pattern,text))   #匹配多个
# print('finditer:',list(re.finditer(pattern,text)))   #匹配多个
# print('sub-repl字符串：',re.sub(pattern,repl,text,count=2))  #替换匹配到的
# print('sub-repl字符串：',re.subn(pattern,repl,text,count=2))  #替换匹配到的,并统计次数

# ########每天定时爬取数据发送到指定邮箱
# import requests,schedule,time,smtplib
# from bs4 import BeautifulSoup
# from email.mime.text import MIMEText
# from email.header import Header
# ##获取每天的天气预报
# def weather():
#     headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'}
#     url='http://www.weather.com.cn/weather/101280601.shtml'
#     huoqu=requests.get(url,headers=headers)
#     # print(huoqu.status_code)
#     huoqu.encoding='utf-8'
#     # print(huoqu.text)
#     jiexi=BeautifulSoup(huoqu.text,'html.parser')
#     tiqus=jiexi.find(class_="t clearfix").find(class_="sky skyid lv1 on")
#     # print(tiqus)
#     wendu=tiqus.find(class_="tem").text
#     tianqi=tiqus.find(class_="wea").text
#     fengli=tiqus.find(class_="win").text
#     wind=tiqus.find(class_="win").find('span')
#     # print(wind['title'])
#     report=('今天温度为：'+str(wendu)+'  '+'天气为：'+tianqi+' '+wind['title']+' '+fengli).replace('\n','')  ##replace去掉爬取数据中的换行符
#     return report
# ##将天气预报发送至指定邮箱
# def mail(report):   ##定义了一个函数，加一个参数
#     mail_host='smtp.isoftstone.com'
#     mail_user='kehana'
#     mail_pass='1QAZ@2WSX'
#     sender='kehana@isoftstone.com'
#     receivers='cainiaoke@163.com'
#     message=MIMEText(report,'plain','utf-8')
#     subject='今日天气预报'
#     message['Subject']=Header(subject,'utf-8')
#     message['From']=sender
#     message['To']=receivers
#     try:
#         smtpObj=smtplib.SMTP()
#         smtpObj.connect(mail_host,25)
#         smtpObj.login(mail_user,mail_pass)
#         smtpObj.sendmail(sender,receivers,message.as_string())
#         smtpObj.quit()
#         print('success')
#     except smtplib.SMTPException as e:
#         print('error',e)
# ##设置定时任务
# def task():
#     print('开始任务')
#     report=weather()
#     mail(report)
#     print('任务完成')
# schedule.every().day.at("15:51").do(task)
# while True:
#     schedule.run_pending()
#     time.sleep(1)

####爬虫同步访问多个网站看下耗时
# import requests,time
# start=time.time()   #设置开始时间   time.time()表示当前时间
# url_list=['https://www.baidu.com/',
#           'https://www.sina.com/',
#           'https://www.qq.com/',
#           'https://www.163.com/',
#           'http://www.iqiyi.com/',
#           'https://www.tmall.com/',
#           'http://www.sohu.com/',
#           'https://www.dogedoge.com/'
#           ]
# for url in url_list:
#     huoqu=requests.get(url)
#     print(url,huoqu.status_code)
# end=time.time()  #设置结束时间
# print(end-start)   #整个程序耗时

####爬虫异步访问多个网站看下耗时
# from gevent import monkey  #从gevent库中导入monkey模块
# monkey.patch_all()    #创建异步任务固定格式
# import gevent,time,requests  #导入gevent
# start = time.time()    #记录开始时间
# url_list = ['https://www.baidu.com/',
#             'https://www.sina.com.cn/',
#             'http://www.sohu.com/',
#             'https://www.qq.com/',
#             'https://www.163.com/',
#             'http://www.iqiyi.com/',
#             'https://www.tmall.com/',
#             'https://www.dogedoge.com/'
#             ]
# def crawler(url):    #定义一个函数，加入url参数
#     r = requests.get(url)   #爬取网站
#     print(url,time.time()-start,r.status_code)   #打印信息
# tasks_list = []   #创建一个空的任务列表
# for url in url_list:
#     task = gevent.spawn(crawler,url)    #用gevent.spawn函数创建爬虫任务
#     tasks_list.append(task)     #将任务加到任务列表
# gevent.joinall(tasks_list)    #执行任务列表中的所有任务，也就是开始爬取每一个网站
# end = time.time()  #记录结束时间
# print(end-start)

#推导式去重一个列表
# old_list=[1,1,1,2,3,4,5,8,6,7,0,0]
# new_list={item for item in old_list}
# print(new_list)

#取出列表中的偶数
# old_list=[0,1,2,3,4,5,6,7,8,9]
# new_list=[item for item in old_list if item % 2 ==0]   # %2==0表示除以2的余数为0，也就是偶数
# print(new_list)

#按分数筛选出学生成绩
# old_student_score={
#     '韩可':{
#         '语文':97,
#         '数学':83,
#         '英语':89
#     },
#     '哈哈':{
#         '语文':95,
#         '数学':89,
#         '英语':80
#     },
#     '呵呵':{
#         '语文':100,
#         '数学':93,
#         '英语':75
#     }
# }
# new_student_score={name:scores for name,scores in old_student_score.items() if scores['数学'] >= 85 }
# print(new_student_score)

#乘法口诀表
# for i in range(1,10):
#     for j in range(1,i+1):
#         print('{}x{}={}'.format(j,i,j*i),end=' ')  #format格式化函数
#     print("")

#windows下在后台运行.py文件     pythonw test.py > test.log
#windows下查看后台进程     tasklist
#windows下杀死后台进程    taskkill /f /im python.exe

# ###爬虫异步访问多个网站看下耗时
# from gevent import monkey  #从gevent库中导入monkey模块
# monkey.patch_all()    #创建异步任务固定格式
# import gevent,time,requests  #导入gevent
# start = time.time()    #记录开始时间
# url_list = ['https://www.baidu.com/',
#             'https://www.sina.com.cn/',
#             'http://www.sohu.com/',
#             'https://www.qq.com/',
#             'https://www.163.com/',
#             'http://www.iqiyi.com/',
#             'https://www.tmall.com/',
#             'https://www.dogedoge.com/'
#             ]
# def crawler(url):    #定义一个函数，加入url参数
#     r = requests.get(url)   #爬取网站
#     print(url,time.time()-start,r.status_code)   #打印信息
# tasks_list = []   #创建一个空的任务列表
# for url in url_list:
#     task = gevent.spawn(crawler,url)    #用gevent.spawn函数创建爬虫任务
#     tasks_list.append(task)     #将任务加到任务列表
# gevent.joinall(tasks_list)    #执行任务列表中的所有任务，也就是开始爬取每一个网站
# end = time.time()  #记录结束时间
# print(end-start)

#爬虫队列
# from gevent import monkey
# monkey.patch_all()
# from gevent.queue import Queue
# import gevent,requests,time
# start=time.time()
# url_list=['https://www.baidu.com/',
#         'https://www.sina.com.cn/',
#         'http://www.sohu.com/',
#         'https://www.qq.com/',
#         'https://www.163.com/',
#         'http://www.iqiyi.com/',
#         'https://www.tmall.com/',
#         'https://www.dogedoge.com/']
# work=Queue()    #创建队列对象work
# for url in url_list:
#     work.put_nowait(url)   #将url放入队列中
# def crawler():    #创建爬虫程序
#     while not work.empty():   #如果队列不为空咋执行下面的步骤
#         url=work.get_nowait()   #从队列中读取url
#         huoqu=requests.get(url)   #requests去请求url
#         print(url,work.qsize(),huoqu.status_code)
# tasks_list=[]   #创建一个空任务列表
# for x in range(2):   #创建两个爬虫
#     task=gevent.spawn(crawler)    #创建爬虫任务
#     tasks_list.append(task)    #将任务加到列表中
# gevent.joinall(tasks_list)   #joinall执行任务列表中的所有任务
# end=time.time()
# print(end-start)
####